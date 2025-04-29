import os
import json
from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import io

class DefectExcelReportGenerator:

    
    def __init__(self, json_data=None):
      
        self.data = json_data
        
            
        self.header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.header_font = Font(bold=True)
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        self.severity_colors = {
            "low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Green
            "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
            "high": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Red
        }
        
    def create_excel(self):
      
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        defect_sheet = workbook.create_sheet("Defect Details")
        
        self._create_summary_sheet(summary_sheet)
        
        self._create_defect_details_sheet(defect_sheet)
        
        workbook.save(buffer)
        buffer.seek(0)  
        return buffer
        
    def _create_summary_sheet(self, sheet):
        """
        Create and populate the summary sheet with model info, charts, and defect counts
        """
     
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        sheet['A1'] = "3D Model Defect Analysis Report"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:E1')
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        sheet['A3'] = "Model:"
        sheet['B3'] = os.path.basename(self.data["model_path"])
        sheet['A3'].font = Font(bold=True)
        
        sheet['A4'] = "Analysis Date:"
        sheet['B4'] = datetime.fromisoformat(self.data["timestamp"]).strftime("%Y-%m-%d %H:%M:%S")
        sheet['A4'].font = Font(bold=True)
        
        sheet['A5'] = "Watertight Model:"
        sheet['B5'] = "Yes" if self.data["watertight"] else "No"
        sheet['B5'].fill = PatternFill(start_color="C6EFCE" if self.data["watertight"] else "FFC7CE", 
                                      end_color="C6EFCE" if self.data["watertight"] else "FFC7CE", 
                                      fill_type="solid")
        sheet['A5'].font = Font(bold=True)
        
        sheet['A6'] = "Total Defects:"
        sheet['B6'] = self.data["defect_count"]
        sheet['B6'].fill = PatternFill(start_color="C6EFCE" if self.data["defect_count"] == 0 else "FFC7CE", 
                                      end_color="C6EFCE" if self.data["defect_count"] == 0 else "FFC7CE", 
                                      fill_type="solid")
        sheet['A6'].font = Font(bold=True)
        
        sheet['A8'] = "Defect Type Summary"
        sheet['A8'].font = Font(size=14, bold=True)
        sheet.merge_cells('A8:C8')
        
        sheet['A9'] = "Defect Type"
        sheet['B9'] = "Count"
        sheet['C9'] = "Percentage"
        for cell in [sheet['A9'], sheet['B9'], sheet['C9']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        defect_types = {}
        severity_counts = {"low": 0, "medium": 0, "high": 0}
        
        for defect in self.data["defects"]:
            defect_type = defect["type"]
            if defect_type not in defect_types:
                defect_types[defect_type] = 0
            defect_types[defect_type] += 1
            
            severity = defect["severity"]
            severity_counts[severity] += 1
        
        row = 10
        for defect_type, count in defect_types.items():
     
            formatted_type = defect_type.replace('_', ' ').title()
            percentage = f"{(count / self.data['defect_count'] * 100):.1f}%" if self.data['defect_count'] > 0 else "0%"
            
            sheet[f'A{row}'] = formatted_type
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = percentage
            
            # Apply styles
            for col in ['A', 'B', 'C']:
                sheet[f'{col}{row}'].border = self.border
                sheet[f'{col}{row}'].alignment = Alignment(horizontal='center' if col != 'A' else 'left')
            
            row += 1
        
        if self.data["defect_count"] > 0:

            type_chart = BarChart()
            type_chart.title = "Defects by Type"
            type_chart.style = 10
            type_chart.x_axis.title = "Defect Type"
            type_chart.y_axis.title = "Count"
            
          
            data = Reference(sheet, min_col=2, min_row=9, max_row=row-1, max_col=2)
            cats = Reference(sheet, min_col=1, min_row=10, max_row=row-1)
            
            type_chart.add_data(data, titles_from_data=True)
            type_chart.set_categories(cats)
            type_chart.shape = 4  # Rectangle shape
            sheet.add_chart(type_chart, "E9")
            

            sheet['A13'] = "Defect Severity Distribution"
            sheet['A13'].font = Font(size=14, bold=True)
            sheet.merge_cells('A13:C13')
            
            sheet['A14'] = "Severity"
            sheet['B14'] = "Count"
            sheet['C14'] = "Percentage"
            for cell in [sheet['A14'], sheet['B14'], sheet['C14']]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.border
            
            severity_row = 15
            for severity, count in severity_counts.items():
                if self.data['defect_count'] > 0:
                    percentage = f"{(count / self.data['defect_count'] * 100):.1f}%"
                else:
                    percentage = "0%"
                
                sheet[f'A{severity_row}'] = severity.title()
                sheet[f'B{severity_row}'] = count
                sheet[f'C{severity_row}'] = percentage
                
                for col in ['A', 'B', 'C']:
                    sheet[f'{col}{severity_row}'].border = self.border
                    sheet[f'{col}{severity_row}'].alignment = Alignment(horizontal='center' if col != 'A' else 'left')
                
              
                sheet[f'A{severity_row}'].fill = self.severity_colors[severity]
                
                severity_row += 1
            
        
        footer_row = row + 20
        sheet[f'A{footer_row}'] = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet[f'A{footer_row}'].font = Font(italic=True)
        sheet.merge_cells(f'A{footer_row}:E{footer_row}')
        sheet[f'A{footer_row}'].alignment = Alignment(horizontal='center')
    
    def _create_defect_details_sheet(self, sheet):

        sheet.column_dimensions['A'].width = 10  # ID
        sheet.column_dimensions['B'].width = 20  # Type
        sheet.column_dimensions['C'].width = 15  # Severity
        sheet.column_dimensions['D'].width = 15  # Location X
        sheet.column_dimensions['E'].width = 15  # Location Y
        sheet.column_dimensions['F'].width = 15  # Location Z
        sheet.column_dimensions['G'].width = 15  # Points Count


        headers = ["ID", "Type", "Severity", "Location X", "Location Y", "Location Z", "Points Count"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Add defect details
        for i, defect in enumerate(self.data["defects"], start=1):
            row = i + 1
            formatted_type = defect["type"].replace('_', ' ').title()
            
            
            sheet.cell(row=row, column=1, value=i)
            sheet.cell(row=row, column=2, value=formatted_type)
            
            severity_cell = sheet.cell(row=row, column=3, value=defect["severity"].upper())
            severity_cell.fill = self.severity_colors[defect["severity"]]
            
            for j, coord in enumerate(defect["location"]):
                sheet.cell(row=row, column=4+j, value=round(coord, 2))
            
            sheet.cell(row=row, column=7, value=defect["points_count"])
            


def generate_excel_from_json(json_data):

    generator = DefectExcelReportGenerator(json_data)
    return generator.create_excel()
